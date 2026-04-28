"""
Aggregate raw screening XLSX files into per-MO JSON summaries.
Algorithm source: raws/алгоритм.docx

Column indices (0-based, header row = row index 1):
  РМЖ: mo=0, date_end=7, refusal=8, expired=9,
        coverage=14 ("+"), precancer cols=20-23 ("+"), zno=26, biopsy=27 ("+")
  КРР: mo=0, date_end=7, refusal=8, expired=9,
        coverage=14 ("+"), neg=15 ("+"), pos=16 ("+"), colonoscopy=18 ("+"),
        precancer=23 ("+"), biopsy=29 ("+"), zno=31
  РШМ: mo=0, date_end=7, refusal=8, expired=9,
        coverage=14 ("+"), precancer cols=21,22,23,30,33 ("+"), zno=39

Run from scripts/ or Webapp/ directory.
Outputs: frontend/public/screening_rmzh.json
         frontend/public/screening_krr.json
         frontend/public/screening_rshm.json
         frontend/public/mos.json
         frontend/public/meta.json
"""
import datetime
import json
import pathlib
import re
import openpyxl

ROOT = pathlib.Path(__file__).parent.parent
RAWS = ROOT / "raws"
OUT  = ROOT / "frontend" / "public"

COORDS_FILE = RAWS / "координаты МО бн.xlsx"
RMJ_FILE    = RAWS / "РМЖ (1)бн.xlsx"
KRR_FILE    = RAWS / "КРР (1)бн.xlsx"
RSHM_FILE   = RAWS / "РШМ (1)бн.xlsx"

# ── helpers ──────────────────────────────────────────────────────────────────

def is_plus(v) -> bool:
    return str(v).strip() == "+"

def parse_date(val):
    """Parse date from openpyxl cell value (datetime object or string)."""
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val
    s = str(val).strip()
    if not s or s in ('None', ''):
        return None
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y.%m.%d'):
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

def core_name(full: str) -> str:
    """Extract the canonical clinic name from quotes, stripping legal prefixes."""
    m = re.search(r'[«""](.+?)[»""]', full)
    if m:
        return m.group(1).strip()
    s = re.sub(r'^(КГП|КДП|ТОО|МУ|ГП)\s+(на\s+ПХВ\s+|на\s+ПХБ\s+)?', '', full, flags=re.I).strip()
    s = re.sub(r'\s+(УЗ\s+ОД|УЗ\s+ОА|управления.+)$', '', s, flags=re.I).strip()
    return s

def best_match(name: str, coord_map: dict) -> tuple[str, float, float] | None:
    """Fuzzy match by core name. Returns (canonical_name, lat, lon) or None."""
    cn = core_name(name).lower()
    for k, v in coord_map.items():
        if core_name(k).lower() == cn:
            return k, v[0], v[1]
    for k, v in coord_map.items():
        ck = core_name(k).lower()
        if cn in ck or ck in cn:
            return k, v[0], v[1]
    return None

# ── load MO coordinates ───────────────────────────────────────────────────────

def load_coords() -> dict[str, tuple[float, float]]:
    wb = openpyxl.load_workbook(COORDS_FILE, read_only=True, data_only=True)
    ws = wb.active
    coord_map: dict[str, tuple[float, float]] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = str(row[0]).strip() if row[0] else ""
        coords_raw = str(row[1]).strip() if row[1] else ""
        if not name or not coords_raw:
            continue
        parts = [p.strip() for p in coords_raw.replace(",", " ").split()]
        nums = []
        for p in parts:
            try:
                nums.append(float(p))
            except ValueError:
                pass
        if len(nums) >= 2:
            coord_map[name] = (nums[0], nums[1])
    wb.close()
    print(f"Loaded {len(coord_map)} MO coordinate entries")
    return coord_map

# ── aggregate one XLSX ────────────────────────────────────────────────────────

def aggregate(path: pathlib.Path, col_spec: dict) -> dict[tuple, dict]:
    """
    Returns dict keyed by (mo_name, year, quarter) with aggregated stats.
    year and quarter may be None if date parsing fails.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    totals: dict[tuple, dict] = {}

    def zero():
        return {"completed": 0, "coverage": 0, "precancers": 0,
                "zno": 0, "refusals": 0, "expired": 0,
                "biopsy": 0, "neg": 0, "pos": 0, "colonoscopy": 0}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= 1:
            continue
        mo = str(row[col_spec["mo"]]).strip() if row[col_spec["mo"]] else ""
        if not mo or mo == "None":
            continue
        date_end_raw = row[col_spec["date_end"]]
        if not date_end_raw or str(date_end_raw).strip() in ("", "None"):
            continue

        d = parse_date(date_end_raw)
        year = d.year if d else None
        month = d.month if d else None
        quarter = ((month - 1) // 3 + 1) if month else None

        key = (mo, year, quarter)
        if key not in totals:
            totals[key] = zero()
        t = totals[key]
        t["completed"] += 1

        if is_plus(row[col_spec["coverage"]]):
            t["coverage"] += 1
        if is_plus(row[col_spec["refusal"]]):
            t["refusals"] += 1
        if is_plus(row[col_spec["expired"]]):
            t["expired"] += 1

        for ci in col_spec.get("precancer", []):
            if is_plus(row[ci]):
                t["precancers"] += 1
                break

        zno_val = row[col_spec["zno"]]
        if zno_val is not None and str(zno_val).strip() not in ("", "None"):
            t["zno"] += 1

        if col_spec.get("biopsy") is not None and is_plus(row[col_spec["biopsy"]]):
            t["biopsy"] += 1
        if col_spec.get("neg") is not None and is_plus(row[col_spec["neg"]]):
            t["neg"] += 1
        if col_spec.get("pos") is not None and is_plus(row[col_spec["pos"]]):
            t["pos"] += 1
        if col_spec.get("colonoscopy") is not None and is_plus(row[col_spec["colonoscopy"]]):
            t["colonoscopy"] += 1

    wb.close()
    return totals

# ── build output JSON ─────────────────────────────────────────────────────────

def build_output(totals: dict[tuple, dict], coord_map: dict, label: str) -> list[dict]:
    # Aggregate by canonical name to collapse duplicates from the same MO
    # (multiple raw rows with different legal prefixes mapping to one canonical name)
    canonical: dict[tuple, dict] = {}
    unmatched: set[str] = set()
    SUM_KEYS = ('completed', 'coverage', 'precancers', 'zno', 'refusals',
                'expired', 'biopsy', 'neg', 'pos', 'colonoscopy')
    for (mo_name, year, quarter), stats in sorted(totals.items()):
        match = best_match(mo_name, coord_map)
        if match is None:
            unmatched.add(mo_name)
            continue
        canonical_name, lat, lon = match
        key = (canonical_name, year, quarter)
        if key not in canonical:
            canonical[key] = {'lat': lat, 'lon': lon, **{k: stats[k] for k in SUM_KEYS}}
        else:
            for k in SUM_KEYS:
                canonical[key][k] += stats[k]

    rows = []
    for (canonical_name, year, quarter), s in sorted(canonical.items()):
        rows.append({
            "mo_name": canonical_name,
            "lat": s['lat'],
            "lon": s['lon'],
            "year": year,
            "quarter": quarter,
            "completed": s["completed"],
            "coverage": s["coverage"],
            "coverage_pct": round(s["coverage"] / s["completed"] * 100, 1) if s["completed"] else 0,
            "precancers": s["precancers"],
            "zno": int(s["zno"]),
            "refusals": s["refusals"],
            "expired": s["expired"],
            "biopsy": s["biopsy"],
            "neg": s["neg"],
            "pos": s["pos"],
            "colonoscopy": s["colonoscopy"],
        })
    if unmatched:
        print(f"  [{label}] {len(unmatched)} MOs not in coordinates file (skipped):")
        for n in sorted(unmatched):
            print(f"    - {n}")
    return rows

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    OUT.mkdir(parents=True, exist_ok=True)
    coord_map = load_coords()

    mos = []
    for name, (lat, lon) in coord_map.items():
        mos.append({"name": name, "lat": lat, "lon": lon})
    (OUT / "mos.json").write_text(json.dumps(mos, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote mos.json ({len(mos)} entries)")

    # РМЖ
    print("\nProcessing РМЖ...")
    rmzh_spec = {
        "mo": 0, "date_end": 7, "refusal": 8, "expired": 9,
        "coverage": 14,
        "precancer": [20, 21, 22, 23],
        "zno": 26, "biopsy": 27,
        "neg": None, "pos": None, "colonoscopy": None,
    }
    rmzh = aggregate(RMJ_FILE, rmzh_spec)
    rmzh_out = build_output(rmzh, coord_map, "РМЖ")
    (OUT / "screening_rmzh.json").write_text(
        json.dumps(rmzh_out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  Wrote screening_rmzh.json ({len(rmzh_out)} records, "
          f"{sum(r['completed'] for r in rmzh_out)} completed)")

    # КРР
    print("\nProcessing КРР...")
    krr_spec = {
        "mo": 0, "date_end": 7, "refusal": 8, "expired": 9,
        "coverage": 14,
        "neg": 15, "pos": 16,
        "colonoscopy": 18,
        "precancer": [23],
        "biopsy": 29,
        "zno": 31,
    }
    krr = aggregate(KRR_FILE, krr_spec)
    krr_out = build_output(krr, coord_map, "КРР")
    (OUT / "screening_krr.json").write_text(
        json.dumps(krr_out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  Wrote screening_krr.json ({len(krr_out)} records, "
          f"{sum(r['completed'] for r in krr_out)} completed)")

    # РШМ
    print("\nProcessing РШМ...")
    rshm_spec = {
        "mo": 0, "date_end": 7, "refusal": 8, "expired": 9,
        "coverage": 14,
        "precancer": [21, 22, 23, 30, 33],
        "zno": 39,
        "biopsy": None, "neg": None, "pos": None, "colonoscopy": None,
    }
    rshm = aggregate(RSHM_FILE, rshm_spec)
    rshm_out = build_output(rshm, coord_map, "РШМ")
    (OUT / "screening_rshm.json").write_text(
        json.dumps(rshm_out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  Wrote screening_rshm.json ({len(rshm_out)} records, "
          f"{sum(r['completed'] for r in rshm_out)} completed)")

    # meta.json — available years and quarters
    all_years: set[int] = set()
    all_quarters: set[int] = set()
    for rec in rmzh_out + krr_out + rshm_out:
        if rec['year'] is not None:
            all_years.add(rec['year'])
        if rec['quarter'] is not None:
            all_quarters.add(rec['quarter'])

    meta = {
        "years": sorted(all_years),
        "quarters": sorted(all_quarters),
    }
    (OUT / "meta.json").write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
    print(f"\nWrote meta.json (years: {sorted(all_years)}, quarters: {sorted(all_quarters)})")
    print("\nDone.")

if __name__ == "__main__":
    main()
