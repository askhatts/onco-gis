"""
Process raws/template_epidemiology-2.xlsx → frontend/public/epidemiology.json

Supported formats:
  Annual  (9 cols):  Район | Год | Заб | Смерт | Смерт/Заб% | Ранняя% | Впервые0-I | Запущ% | Выжив%
  Quarterly (10 cols): Район | Год | Квартал | Заб | Смерт | Смерт/Заб% | Ранняя% | Впервые0-I | Запущ% | Выжив%

Annual records are automatically expanded to Q1–Q4 with identical values.
"""
import json
import pathlib
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = pathlib.Path(__file__).parent.parent
RAWS = ROOT / "raws"
OUT  = ROOT / "frontend" / "public"

EPI_FILE = RAWS / "template_epidemiology-2.xlsx"
TEMPLATE_QUARTERLY = RAWS / "template_epidemiology-quarterly.xlsx"

INDICATORS = [
    {"id": "incidence_rate",     "label": "Заболеваемость",            "unit": "на 100 тыс.", "color": "#00c4ce", "polarity": "negative"},
    {"id": "mortality_rate",     "label": "Смертность",                "unit": "на 100 тыс.", "color": "#e85050", "polarity": "negative"},
    {"id": "mortality_ratio",    "label": "Смертность/заболеваемость", "unit": "%",           "color": "#e8a020", "polarity": "negative"},
    {"id": "early_stage_pct",    "label": "Ранняя диагностика",        "unit": "%",           "color": "#27c97a", "polarity": "positive"},
    {"id": "early_stage_count",  "label": "Впервые 0–I ст.",           "unit": "чел.",        "color": "#00a89e", "polarity": "positive"},
    {"id": "advanced_stage_pct", "label": "Запущенность",              "unit": "%",           "color": "#e85050", "polarity": "negative"},
    {"id": "survival_5yr_pct",   "label": "5-лет. выживаемость",       "unit": "%",           "color": "#3a8ff4", "polarity": "positive"},
]

ABAY_DISTRICTS = [
    "г. Семей", "Абайский район", "Аксуатский район", "Аягозский район",
    "Бескарагайский район", "Бородулихинский район", "Кокпектинский район",
    "г. Курчатов", "Урджарский район", "Жарминский район",
    "Жанасемейский район", "Маканчинский район",
]

HEADERS_QUARTERLY = [
    "Район", "Год", "Квартал",
    "Заболеваемость", "Смертность", "Смертность_к_заболеваемости_%",
    "Ранняя_диагностика_%", "число впервые выявленных больных ЗН 0-I ст",
    "Запущенность_%", "Выживаемость_5лет_%",
]


def generate_quarterly_template(path: pathlib.Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Эпидемиология"
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    for col, header in enumerate(HEADERS_QUARTERLY, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    for col_letter in 'DEFGHIJ':
        ws.column_dimensions[col_letter].width = 20
    row = 2
    for district in ABAY_DISTRICTS:
        for year in [2025, 2026]:
            for quarter in [1, 2, 3, 4]:
                ws.cell(row=row, column=1, value=district)
                ws.cell(row=row, column=2, value=year)
                ws.cell(row=row, column=3, value=quarter)
                row += 1
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Generated quarterly template: {path}")


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def is_quarterly_format(header_row) -> bool:
    """Detect if file has Квартал column (10+ cols, col[2] header contains 'квартал')."""
    if not header_row or len(header_row) < 10:
        return False
    return 'кварт' in str(header_row[2] or '').lower()


def process_epidemiology(path: pathlib.Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    quarterly = is_quarterly_format(rows[0])
    data = []

    for i, row in enumerate(rows):
        if i == 0:
            continue
        district = str(row[0]).strip() if row[0] else ""
        if not district or district == "None":
            continue
        try:
            year = int(row[1]) if row[1] is not None else None
        except (ValueError, TypeError):
            year = None

        if quarterly:
            try:
                quarter = int(row[2]) if row[2] is not None else None
            except (ValueError, TypeError):
                quarter = None
            ind_offset = 3
        else:
            quarter = None
            ind_offset = 2

        def col(i):
            idx = ind_offset + i
            return to_float(row[idx] if len(row) > idx else None)

        base = {
            "district_name_ru":  district,
            "year":              year,
            "incidence_rate":    col(0),
            "mortality_rate":    col(1),
            "mortality_ratio":   col(2),
            "early_stage_pct":   col(3),
            "early_stage_count": col(4),
            "advanced_stage_pct":col(5),
            "survival_5yr_pct":  col(6),
        }

        if quarter is not None:
            base["quarter"] = quarter
            data.append(base)
        else:
            # Expand annual record to all 4 quarters
            for q in [1, 2, 3, 4]:
                data.append({**base, "quarter": q})

    return data


def main():
    OUT.mkdir(parents=True, exist_ok=True)

    if not TEMPLATE_QUARTERLY.exists():
        generate_quarterly_template(TEMPLATE_QUARTERLY)

    if EPI_FILE.exists():
        data = process_epidemiology(EPI_FILE)
        print(f"Processed {len(data)} epidemiology records from {EPI_FILE.name}")
    else:
        print(f"{EPI_FILE.name} not found in raws/ — writing empty data")
        data = []

    output = {"indicators": INDICATORS, "data": data}
    (OUT / "epidemiology.json").write_text(
        json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Wrote epidemiology.json ({len(data)} records)")

    # Merge epi years into meta.json
    meta_path = OUT / "meta.json"
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        meta = {"years": [], "quarters": [1, 2, 3, 4]}
    epi_years = {r["year"] for r in data if r.get("year") is not None}
    merged_years = sorted(set(meta.get("years", [])) | epi_years)
    meta["years"] = merged_years
    meta_path.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
    print(f"Updated meta.json years: {merged_years}")


if __name__ == "__main__":
    main()
